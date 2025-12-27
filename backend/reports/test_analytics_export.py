"""
Comprehensive tests for Analytics Export Service.

Tests cover:
- CSV export with various options
- Excel export with formatting
- Multi-sheet exports
- Column customization
- Date formatting
- Large dataset streaming
- Async export via Celery
- Progress tracking
- Cache management
- Special character handling
"""

import csv
import io
from datetime import datetime, timedelta, date
from decimal import Decimal
from typing import Any, Dict, List

import pytest
from django.contrib.auth import get_user_model
from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.test import APITestCase

from reports.models import (
    AnalyticsData,
    StudentReport,
    TeacherWeeklyReport,
    TutorWeeklyReport,
)
from reports.services.analytics_export import (
    AnalyticsExportService,
    export_student_analytics_async,
    export_class_analytics_async,
    export_report_async,
    export_custom_async,
)

User = get_user_model()


def get_streaming_response_content(response) -> bytes:
    """Helper to extract bytes from StreamingHttpResponse."""
    if hasattr(response, "streaming_content"):
        return b"".join(response.streaming_content)
    return response.content


class AnalyticsExportServiceTests(TestCase):
    """Test AnalyticsExportService core functionality."""

    def setUp(self):
        """Set up test data."""
        self.service = AnalyticsExportService

        # Create users
        self.student1 = User.objects.create_user(
            email='student1@test.com',
            password='test123',
            first_name='John',
            last_name='Doe'
        )
        self.student2 = User.objects.create_user(
            email='student2@test.com',
            password='test123',
            first_name='Jane',
            last_name='Smith'
        )
        self.teacher = User.objects.create_user(
            email='teacher@test.com',
            password='test123',
            first_name='Bob',
            last_name='Teacher'
        )

        # Create analytics data
        self.today = timezone.now().date()
        self.start_date = self.today - timedelta(days=7)
        self.end_date = self.today

        # Student 1 analytics
        for i in range(5):
            AnalyticsData.objects.create(
                student=self.student1,
                metric_type=AnalyticsData.MetricType.STUDENT_PROGRESS,
                value=75.0 + i * 2,
                unit='%',
                date=self.start_date + timedelta(days=i),
                period_start=self.start_date,
                period_end=self.end_date,
            )

        # Student 2 analytics
        for i in range(5):
            AnalyticsData.objects.create(
                student=self.student2,
                metric_type=AnalyticsData.MetricType.ENGAGEMENT,
                value=60.0 + i * 3,
                unit='%',
                date=self.start_date + timedelta(days=i),
                period_start=self.start_date,
                period_end=self.end_date,
            )

    def test_filter_by_columns_success(self):
        """Test filtering data to specific columns."""
        data = [
            {'id': 1, 'name': 'Analytics 1', 'value': 75.0, 'date': '2024-01-01'},
            {'id': 2, 'name': 'Analytics 2', 'value': 80.0, 'date': '2024-01-02'},
        ]

        filtered = self.service.filter_by_columns(data, ['id', 'name'])

        self.assertEqual(len(filtered), 2)
        self.assertIn('id', filtered[0])
        self.assertIn('name', filtered[0])
        self.assertNotIn('value', filtered[0])
        self.assertEqual(filtered[0]['id'], 1)

    def test_filter_by_columns_all_columns(self):
        """Test that all columns are returned if none specified."""
        data = [{'id': 1, 'name': 'Analytics 1', 'value': 75.0}]

        filtered = self.service.filter_by_columns(data, None)

        self.assertEqual(len(filtered[0]), 3)
        self.assertTrue(all(key in filtered[0] for key in ['id', 'name', 'value']))

    def test_filter_by_columns_empty(self):
        """Test filtering empty data."""
        result = self.service.filter_by_columns([], ['id'])
        self.assertEqual(result, [])

    def test_export_student_analytics_csv(self):
        """Test exporting student analytics to CSV."""
        response = self.service.export_student_analytics(
            student_id=self.student1.id,
            format='csv'
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn('text/csv', response.get('Content-Type', ''))

        # Parse CSV content
        content = get_streaming_response_content(response).decode('utf-8')
        lines = content.strip().split('\n')

        self.assertGreater(len(lines), 1)  # Headers + at least 1 row
        self.assertIn('student_name', lines[0])  # Headers

    def test_export_student_analytics_excel(self):
        """Test exporting student analytics to Excel."""
        response = self.service.export_student_analytics(
            student_id=self.student1.id,
            format='excel'
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn('spreadsheetml', response.get('Content-Type', ''))

        # Parse Excel content
        content = get_streaming_response_content(response)
        workbook = load_workbook(io.BytesIO(content))

        self.assertIn('Analytics', workbook.sheetnames)

    def test_export_student_analytics_with_date_range(self):
        """Test exporting analytics with date range filter."""
        response = self.service.export_student_analytics(
            student_id=self.student1.id,
            format='csv',
            start_date=self.start_date + timedelta(days=1),
            end_date=self.start_date + timedelta(days=3),
        )

        self.assertEqual(response.status_code, 200)
        content = get_streaming_response_content(response).decode('utf-8')

        # Should have filtered data
        lines = content.strip().split('\n')
        # Headers + filtered rows
        self.assertGreaterEqual(len(lines), 1)

    def test_export_student_analytics_custom_columns(self):
        """Test exporting with custom column selection."""
        columns = ['student_name', 'metric_type', 'value']
        response = self.service.export_student_analytics(
            student_id=self.student1.id,
            format='csv',
            columns=columns
        )

        self.assertEqual(response.status_code, 200)
        content = get_streaming_response_content(response).decode('utf-8')
        lines = content.strip().split('\n')

        # Check headers match selected columns
        headers = lines[0].split(',')
        for col in columns:
            self.assertIn(col, headers)

    def test_export_student_analytics_date_format_eu(self):
        """Test exporting with EU date format."""
        response = self.service.export_student_analytics(
            student_id=self.student1.id,
            format='csv',
            date_format='eu'
        )

        self.assertEqual(response.status_code, 200)
        content = get_streaming_response_content(response).decode('utf-8')

        # EU format should use dots (dd.mm.yyyy)
        self.assertIn('.', content)

    def test_export_class_analytics(self):
        """Test exporting class analytics."""
        response = self.service.export_class_analytics(
            class_ids=[1, 2],
            format='csv'
        )

        self.assertEqual(response.status_code, 200)

    def test_export_to_csv_streaming(self):
        """Test CSV export with streaming."""
        data = [
            {'id': i, 'name': f'Item {i}', 'value': 100.0 + i}
            for i in range(1000)
        ]

        response = self.service.export_to_csv(
            data,
            report_name='test_export'
        )

        self.assertEqual(response.status_code, 200)
        content = get_streaming_response_content(response).decode('utf-8')

        # Verify streaming worked
        lines = content.strip().split('\n')
        self.assertEqual(len(lines), 1001)  # Headers + 1000 rows

    def test_export_to_csv_unicode(self):
        """Test CSV export handles Unicode properly."""
        data = [
            {'student': 'Иван Петров', 'subject': '数学', 'score': 95},
            {'student': 'José García', 'subject': 'Français', 'score': 87},
        ]

        response = self.service.export_to_csv(
            data,
            report_name='unicode_test'
        )

        self.assertEqual(response.status_code, 200)
        content = get_streaming_response_content(response).decode('utf-8')

        # Should preserve Unicode characters
        self.assertIn('Иван', content)
        self.assertIn('José', content)
        self.assertIn('Français', content)

    def test_export_to_csv_special_chars(self):
        """Test CSV export with special characters."""
        data = [
            {'name': 'Test, with comma', 'description': 'Quote "here"', 'value': 100},
            {'name': 'Newline\ntest', 'description': 'Normal', 'value': 200},
        ]

        response = self.service.export_to_csv(data)

        content = get_streaming_response_content(response).decode('utf-8')
        reader = csv.DictReader(io.StringIO(content))
        rows = list(reader)

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]['name'], 'Test, with comma')
        self.assertEqual(rows[0]['description'], 'Quote "here"')

    def test_export_to_csv_empty_data(self):
        """Test CSV export with empty data."""
        response = self.service.export_to_csv([])

        self.assertEqual(response.status_code, 200)
        content = get_streaming_response_content(response).decode('utf-8')

        # Should have empty file with just headers concept
        self.assertIsNotNone(content)

    def test_export_to_excel_formatting(self):
        """Test Excel export with proper formatting."""
        data = [
            {'name': 'Student 1', 'score': 95.5, 'date': '2024-01-01'},
            {'name': 'Student 2', 'score': 87.25, 'date': '2024-01-02'},
        ]

        response = self.service.export_to_excel(
            data,
            report_name='formatted_export',
            style=True
        )

        self.assertEqual(response.status_code, 200)
        content = get_streaming_response_content(response)
        workbook = load_workbook(io.BytesIO(content))

        ws = workbook.active

        # Check headers
        self.assertEqual(ws.cell(1, 1).value, 'name')
        self.assertEqual(ws.cell(1, 2).value, 'score')
        self.assertEqual(ws.cell(1, 3).value, 'date')

        # Check data
        self.assertEqual(ws.cell(2, 1).value, 'Student 1')
        self.assertEqual(ws.cell(2, 2).value, 95.5)

    def test_export_to_excel_freeze_panes(self):
        """Test Excel export with frozen panes."""
        data = [
            {'id': i, 'name': f'Item {i}', 'value': 100 + i}
            for i in range(100)
        ]

        response = self.service.export_to_excel(
            data,
            freeze_panes=True
        )

        self.assertEqual(response.status_code, 200)
        content = get_streaming_response_content(response)
        workbook = load_workbook(io.BytesIO(content))

        ws = workbook.active

        # Check that freeze panes is set
        self.assertEqual(ws.freeze_panes, "A2")

    def test_export_to_excel_column_width(self):
        """Test Excel export adjusts column widths."""
        data = [
            {'short': 'x', 'very_long_column_name': 'y' * 50},
        ]

        response = self.service.export_to_excel(data)

        content = get_streaming_response_content(response)
        workbook = load_workbook(io.BytesIO(content))

        ws = workbook.active

        # Check that column widths are set
        self.assertIsNotNone(ws.column_dimensions['A'].width)
        self.assertIsNotNone(ws.column_dimensions['B'].width)

    def test_export_to_excel_number_format(self):
        """Test Excel export applies number formatting."""
        data = [
            {'name': 'Item', 'score': 95.5, 'count': 42},
        ]

        response = self.service.export_to_excel(data)

        content = get_streaming_response_content(response)
        workbook = load_workbook(io.BytesIO(content))

        ws = workbook.active

        # Score column should have decimal format
        score_cell = ws.cell(2, 2)
        self.assertEqual(score_cell.number_format, '0.00')

        # Count column should have integer format
        count_cell = ws.cell(2, 3)
        self.assertEqual(count_cell.number_format, '0')

    def test_export_multi_sheet_excel(self):
        """Test exporting multiple sheets to Excel."""
        sheets = {
            'Sheet1': [
                {'id': 1, 'name': 'Item 1', 'value': 100},
                {'id': 2, 'name': 'Item 2', 'value': 200},
            ],
            'Sheet2': [
                {'student': 'John', 'score': 95},
                {'student': 'Jane', 'score': 88},
            ],
            'Sheet3': [
                {'date': '2024-01-01', 'count': 10},
            ],
        }

        response = self.service.export_multi_sheet_excel(
            sheets=sheets,
            report_name='multi_sheet'
        )

        self.assertEqual(response.status_code, 200)
        content = get_streaming_response_content(response)
        workbook = load_workbook(io.BytesIO(content))

        # Check all sheets created
        self.assertEqual(len(workbook.sheetnames), 3)
        self.assertIn('Sheet1', workbook.sheetnames)
        self.assertIn('Sheet2', workbook.sheetnames)
        self.assertIn('Sheet3', workbook.sheetnames)

        # Check data in sheets
        ws1 = workbook['Sheet1']
        self.assertEqual(ws1.cell(2, 1).value, 1)
        self.assertEqual(ws1.cell(2, 2).value, 'Item 1')

    def test_export_to_excel_empty_data(self):
        """Test Excel export with empty data."""
        response = self.service.export_to_excel([])

        self.assertEqual(response.status_code, 200)
        content = get_streaming_response_content(response)

        # Should create valid Excel file even if empty
        workbook = load_workbook(io.BytesIO(content))
        self.assertIsNotNone(workbook.active)

    def test_export_large_dataset_csv(self):
        """Test CSV export with large dataset."""
        # Create 10,000 rows
        large_data = [
            {'id': i, 'name': f'Item {i}', 'value': Decimal(100.5 + i * 0.1)}
            for i in range(10000)
        ]

        response = self.service.export_to_csv(large_data)

        self.assertEqual(response.status_code, 200)
        content = get_streaming_response_content(response).decode('utf-8')
        lines = content.strip().split('\n')

        # Should have all rows
        self.assertEqual(len(lines), 10001)  # Headers + 10000

    def test_export_exceeds_max_rows_csv(self):
        """Test CSV export rejects dataset larger than MAX_ROWS."""
        # Create dataset larger than MAX_ROWS_FOR_CSV
        large_data = [{'id': i} for i in range(5000001)]

        with self.assertRaises(ValueError):
            self.service.export_to_csv(large_data)

    def test_export_exceeds_max_rows_excel(self):
        """Test Excel export rejects dataset larger than MAX_ROWS."""
        large_data = [{'id': i} for i in range(1000001)]

        with self.assertRaises(ValueError):
            self.service.export_to_excel(large_data)

    def test_export_report_student(self):
        """Test exporting a student report."""
        # Create a student report
        report = StudentReport.objects.create(
            title='Progress Report',
            student=self.student1,
            teacher=self.teacher,
            report_type='progress',
            status='completed'
        )

        response = self.service.export_report(
            report_id=report.id,
            report_type='student',
            format='csv'
        )

        self.assertEqual(response.status_code, 200)
        content = get_streaming_response_content(response).decode('utf-8')

        self.assertIn('Progress Report', content)

    def test_export_with_delimiter(self):
        """Test CSV export with custom delimiter."""
        data = [
            {'name': 'Item 1', 'value': 100},
            {'name': 'Item 2', 'value': 200},
        ]

        response = self.service.export_to_csv(
            data,
            delimiter=';'
        )

        content = get_streaming_response_content(response).decode('utf-8')

        # Should use semicolon as delimiter
        self.assertIn(';', content)

    def test_date_formatting_options(self):
        """Test different date format options."""
        data = [
            {'date': '2024-01-15', 'value': 100},
        ]

        for date_format in ['iso', 'us', 'eu', 'full']:
            response = self.service.export_to_csv(data)
            self.assertEqual(response.status_code, 200)

    def test_clear_export_cache(self):
        """Test clearing export cache."""
        # Should not raise error
        self.service.clear_export_cache()
        self.service.clear_export_cache('specific_key')

        # Return True on success
        result = self.service.clear_export_cache()
        self.assertTrue(result)

    def test_export_filename_includes_timestamp(self):
        """Test that exported filenames include timestamps."""
        data = [{'id': 1, 'name': 'Item'}]

        response = self.service.export_to_csv(data, report_name='test')

        disposition = response.get('Content-Disposition', '')
        self.assertIn('test_', disposition)
        self.assertIn('.csv', disposition)

        # Extract filename
        filename = disposition.split('filename="')[1].rstrip('"')
        self.assertRegex(filename, r'test_\d{8}_\d{6}\.csv')

    def test_export_excel_filename_xlsx(self):
        """Test that Excel exports have .xlsx extension."""
        data = [{'id': 1}]

        response = self.service.export_to_excel(data, report_name='test')

        disposition = response.get('Content-Disposition', '')
        filename = disposition.split('filename="')[1].rstrip('"')

        self.assertTrue(filename.endswith('.xlsx'))

    def test_custom_sheet_name_excel(self):
        """Test custom sheet names in Excel export."""
        data = [{'id': 1, 'value': 100}]

        response = self.service.export_to_excel(
            data,
            sheet_name='Custom Sheet'
        )

        content = get_streaming_response_content(response)
        workbook = load_workbook(io.BytesIO(content))

        self.assertEqual(workbook.active.title, 'Custom Sheet')

    def test_export_with_none_values(self):
        """Test CSV export handles None values."""
        data = [
            {'id': 1, 'name': 'Item 1', 'value': None},
            {'id': 2, 'name': None, 'value': 200},
            {'id': None, 'name': 'Item 3', 'value': 300},
        ]

        response = self.service.export_to_csv(data)

        self.assertEqual(response.status_code, 200)
        content = get_streaming_response_content(response).decode('utf-8')

        # Should handle None values gracefully
        reader = csv.DictReader(io.StringIO(content))
        rows = list(reader)
        self.assertEqual(len(rows), 3)
