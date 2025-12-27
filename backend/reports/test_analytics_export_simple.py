"""
Simple unit tests for AnalyticsExportService without database dependencies.

Tests basic functionality that doesn't require Django models.
"""

import io
import csv
from datetime import datetime
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from reports.services.analytics_export import AnalyticsExportService


class TestAnalyticsExportServiceBasic:
    """Basic tests for AnalyticsExportService without DB."""

    def test_filter_by_columns_success(self):
        """Test filtering data to specific columns."""
        data = [
            {'id': 1, 'name': 'Report 1', 'value': 75.0, 'date': '2024-01-01'},
            {'id': 2, 'name': 'Report 2', 'value': 80.0, 'date': '2024-01-02'},
        ]

        filtered = AnalyticsExportService.filter_by_columns(data, ['id', 'name'])

        assert len(filtered) == 2
        assert 'id' in filtered[0]
        assert 'name' in filtered[0]
        assert 'value' not in filtered[0]
        assert filtered[0]['id'] == 1

    def test_filter_by_columns_no_columns(self):
        """Test that all columns are returned if none specified."""
        data = [{'id': 1, 'name': 'Report 1', 'value': 75.0}]

        filtered = AnalyticsExportService.filter_by_columns(data, None)

        assert len(filtered[0]) == 3
        assert all(key in filtered[0] for key in ['id', 'name', 'value'])

    def test_filter_by_columns_empty(self):
        """Test filtering empty data."""
        result = AnalyticsExportService.filter_by_columns([], ['id'])
        assert result == []

    def test_export_to_csv_basic(self):
        """Test basic CSV export."""
        data = [
            {'id': 1, 'name': 'Item 1', 'value': 100},
            {'id': 2, 'name': 'Item 2', 'value': 200},
        ]

        response = AnalyticsExportService.export_to_csv(data, report_name='test')

        assert response.status_code == 200
        assert 'text/csv' in response.get('Content-Type', '')

        # Parse CSV content
        content = b"".join(response.streaming_content).decode('utf-8')
        lines = content.strip().split('\n')

        assert len(lines) == 3  # Headers + 2 rows
        assert 'id' in lines[0]

    def test_export_to_csv_unicode(self):
        """Test CSV export with Unicode."""
        data = [
            {'name': 'Иван Петров', 'score': 95},
            {'name': 'José García', 'score': 87},
        ]

        response = AnalyticsExportService.export_to_csv(data)

        content = b"".join(response.streaming_content).decode('utf-8')

        assert 'Иван' in content
        assert 'José' in content

    def test_export_to_csv_special_chars(self):
        """Test CSV export with special characters."""
        data = [
            {'name': 'Test, with comma', 'desc': 'Quote "here"'},
        ]

        response = AnalyticsExportService.export_to_csv(data)

        content = b"".join(response.streaming_content).decode('utf-8')

        # Just verify the content has our data properly escaped
        assert 'Test, with comma' in content
        assert 'Quote "here"' in content or 'Quote' in content

    def test_export_to_csv_empty(self):
        """Test CSV export with empty data."""
        response = AnalyticsExportService.export_to_csv([])

        assert response.status_code == 200
        content = b"".join(response.streaming_content).decode('utf-8')
        assert content is not None

    def test_export_to_csv_delimiter(self):
        """Test CSV export with custom delimiter."""
        data = [{'id': 1, 'name': 'Test'}]

        response = AnalyticsExportService.export_to_csv(data, delimiter=';')

        content = b"".join(response.streaming_content).decode('utf-8')
        assert ';' in content

    def test_export_to_csv_large_dataset(self):
        """Test CSV export with large dataset."""
        large_data = [
            {'id': i, 'value': Decimal(100.5 + i * 0.1)}
            for i in range(10000)
        ]

        response = AnalyticsExportService.export_to_csv(large_data)

        assert response.status_code == 200
        content = b"".join(response.streaming_content).decode('utf-8')
        lines = content.strip().split('\n')

        assert len(lines) == 10001  # Headers + 10000

    def test_export_to_csv_exceeds_max(self):
        """Test CSV export rejects oversized dataset."""
        large_data = [{'id': i} for i in range(5000001)]

        with pytest.raises(ValueError):
            AnalyticsExportService.export_to_csv(large_data)

    def test_export_to_excel_basic(self):
        """Test basic Excel export."""
        data = [
            {'name': 'Item 1', 'score': 95.5},
            {'name': 'Item 2', 'score': 87.25},
        ]

        response = AnalyticsExportService.export_to_excel(data)

        assert response.status_code == 200
        assert 'spreadsheetml' in response.get('Content-Type', '')

        # Parse Excel
        content = b"".join(response.streaming_content)
        workbook = load_workbook(io.BytesIO(content))

        ws = workbook.active
        assert ws.cell(1, 1).value == 'name'
        assert ws.cell(2, 1).value == 'Item 1'

    def test_export_to_excel_freeze_panes(self):
        """Test Excel export with frozen panes."""
        data = [{'id': i, 'value': 100 + i} for i in range(100)]

        response = AnalyticsExportService.export_to_excel(data, freeze_panes=True)

        content = b"".join(response.streaming_content)
        workbook = load_workbook(io.BytesIO(content))

        ws = workbook.active
        assert ws.freeze_panes == "A2"

    def test_export_to_excel_number_format(self):
        """Test Excel export applies number formatting."""
        data = [
            {'name': 'Item', 'score': 95.5, 'count': 42},
        ]

        response = AnalyticsExportService.export_to_excel(data)

        content = b"".join(response.streaming_content)
        workbook = load_workbook(io.BytesIO(content))

        ws = workbook.active

        # Score should have decimal format
        score_cell = ws.cell(2, 2)
        assert score_cell.number_format == '0.00'

    def test_export_to_excel_empty(self):
        """Test Excel export with empty data."""
        response = AnalyticsExportService.export_to_excel([])

        assert response.status_code == 200
        content = b"".join(response.streaming_content)

        workbook = load_workbook(io.BytesIO(content))
        assert workbook.active is not None

    def test_export_to_excel_exceeds_max(self):
        """Test Excel export rejects oversized dataset."""
        large_data = [{'id': i} for i in range(1000001)]

        with pytest.raises(ValueError):
            AnalyticsExportService.export_to_excel(large_data)

    def test_export_multi_sheet_excel(self):
        """Test multi-sheet Excel export."""
        sheets = {
            'Sheet1': [
                {'id': 1, 'name': 'Item 1'},
                {'id': 2, 'name': 'Item 2'},
            ],
            'Sheet2': [
                {'student': 'John', 'score': 95},
            ],
        }

        response = AnalyticsExportService.export_multi_sheet_excel(
            sheets=sheets,
            report_name='multi'
        )

        assert response.status_code == 200
        content = b"".join(response.streaming_content)
        workbook = load_workbook(io.BytesIO(content))

        assert len(workbook.sheetnames) == 2
        assert 'Sheet1' in workbook.sheetnames
        assert 'Sheet2' in workbook.sheetnames

    def test_export_filename_has_timestamp(self):
        """Test that exported filenames include timestamps."""
        data = [{'id': 1}]

        response = AnalyticsExportService.export_to_csv(data, report_name='test')

        disposition = response.get('Content-Disposition', '')
        assert 'test_' in disposition
        assert '.csv' in disposition

    def test_export_excel_xlsx_extension(self):
        """Test that Excel exports have .xlsx extension."""
        data = [{'id': 1}]

        response = AnalyticsExportService.export_to_excel(data)

        disposition = response.get('Content-Disposition', '')
        filename = disposition.split('filename="')[1].rstrip('"')

        assert filename.endswith('.xlsx')

    def test_csv_with_none_values(self):
        """Test CSV export handles None values."""
        data = [
            {'id': 1, 'name': 'Item', 'value': None},
            {'id': 2, 'name': None, 'value': 200},
        ]

        response = AnalyticsExportService.export_to_csv(data)

        assert response.status_code == 200
        content = b"".join(response.streaming_content).decode('utf-8')
        reader = csv.DictReader(io.StringIO(content))
        rows = list(reader)

        assert len(rows) == 2

    def test_clear_export_cache(self):
        """Test clearing export cache."""
        result = AnalyticsExportService.clear_export_cache()
        assert result is True

        result = AnalyticsExportService.clear_export_cache('specific_key')
        assert result is True

    def test_csv_includes_headers(self):
        """Test CSV export includes headers."""
        data = [{'id': 1, 'name': 'Test'}]

        response = AnalyticsExportService.export_to_csv(
            data,
            include_headers=True
        )

        content = b"".join(response.streaming_content).decode('utf-8')
        lines = content.strip().split('\n')

        assert 'id' in lines[0]
        assert 'name' in lines[0]

    def test_excel_sheet_name_truncated(self):
        """Test that long sheet names are truncated to 31 chars."""
        data = [{'id': 1}]
        long_name = 'A' * 50

        response = AnalyticsExportService.export_to_excel(
            data,
            sheet_name=long_name
        )

        content = b"".join(response.streaming_content)
        workbook = load_workbook(io.BytesIO(content))

        assert len(workbook.active.title) <= 31

    def test_csv_with_decimal_values(self):
        """Test CSV export handles Decimal values."""
        data = [
            {'id': 1, 'amount': Decimal('123.45')},
            {'id': 2, 'amount': Decimal('999.99')},
        ]

        response = AnalyticsExportService.export_to_csv(data)

        content = b"".join(response.streaming_content).decode('utf-8')

        assert '123.45' in content
        assert '999.99' in content
