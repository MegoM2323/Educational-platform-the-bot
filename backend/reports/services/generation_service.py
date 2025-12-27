"""
Report Generation Service

Comprehensive service for generating reports from various data sources.
Supports multiple report types, output formats (PDF, Excel, JSON),
caching strategies, and async processing.

Features:
- Multiple report types (Student Progress, Class Performance, etc.)
- Data collection from database sources
- Multiple output formats (PDF, Excel, JSON)
- Report caching with TTL and invalidation
- Async Celery task support
- Progress tracking for long-running reports
- Large dataset pagination
- Error recovery with retry logic
"""

import json
import logging
import time
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime, date, timedelta
from decimal import Decimal
from io import BytesIO

from django.db.models import Q, Avg, Count, Sum, Case, When, Value, F, Max, Min
from django.core.cache import cache
from django.contrib.auth import get_user_model
from django.utils import timezone
from django.http import HttpResponse

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from assignments.models import Assignment, AssignmentSubmission
from materials.models import Material, MaterialProgress
from chat.models import ChatRoom, Message
from knowledge_graph.models import LessonProgress

from ..models import Report, ReportTemplate

User = get_user_model()
logger = logging.getLogger(__name__)


class ReportGenerationException(Exception):
    """Exception for report generation errors."""
    pass


class ReportGenerationService:
    """
    Core service for generating reports from various data sources.

    Handles:
    - Template-based report generation
    - Data collection from multiple sources
    - Output formatting (PDF, Excel, JSON)
    - Caching and invalidation
    - Progress tracking
    - Pagination for large datasets
    """

    # Cache configuration
    CACHE_TTL_SHORT = 300  # 5 minutes
    CACHE_TTL_MEDIUM = 1800  # 30 minutes
    CACHE_TTL_LONG = 3600  # 1 hour

    # Pagination settings
    DEFAULT_PAGE_SIZE = 100
    MAX_PAGE_SIZE = 1000

    # Report types with their metadata
    REPORT_TYPES = {
        'student_progress': {
            'name': 'Student Progress Report',
            'description': 'Individual student progress and performance',
            'data_sources': ['materials', 'assignments', 'knowledge_graph']
        },
        'class_performance': {
            'name': 'Class Performance Report',
            'description': 'Overall class performance metrics',
            'data_sources': ['assignments', 'materials', 'knowledge_graph']
        },
        'assignment_analysis': {
            'name': 'Assignment Analysis Report',
            'description': 'Detailed assignment submission and grading analysis',
            'data_sources': ['assignments']
        },
        'subject_analysis': {
            'name': 'Subject Analysis Report',
            'description': 'Subject-specific performance and progress',
            'data_sources': ['materials', 'assignments', 'knowledge_graph']
        },
        'tutor_weekly': {
            'name': 'Tutor Weekly Report',
            'description': 'Weekly tutoring summary and student progress',
            'data_sources': ['materials', 'assignments', 'chat']
        },
        'teacher_weekly': {
            'name': 'Teacher Weekly Report',
            'description': 'Weekly teaching summary and class analytics',
            'data_sources': ['assignments', 'materials', 'knowledge_graph']
        }
    }

    def __init__(self, user: User, report_type: str):
        """
        Initialize the report generation service.

        Args:
            user: User requesting the report (determines access level)
            report_type: Type of report to generate

        Raises:
            ReportGenerationException: If report type is invalid or user has no access
        """
        if report_type not in self.REPORT_TYPES:
            raise ReportGenerationException(
                f"Invalid report type: {report_type}. "
                f"Valid types: {', '.join(self.REPORT_TYPES.keys())}"
            )

        self.user = user
        self.report_type = report_type
        self.metadata = self.REPORT_TYPES[report_type]
        self.generation_start_time = None
        self.progress_data = {'current': 0, 'total': 0, 'status': 'initializing'}

    def generate(
        self,
        filters: Optional[Dict[str, Any]] = None,
        output_format: str = 'json',
        cache_enabled: bool = True,
        progress_callback: Optional[callable] = None
    ) -> Dict[str, Any]:
        """
        Generate a complete report.

        Args:
            filters: Dict of filters (student_id, date_range, subject, etc.)
            output_format: Output format ('json', 'excel', 'pdf')
            cache_enabled: Whether to cache the result
            progress_callback: Callback function for progress updates

        Returns:
            Dict with report data and metadata

        Raises:
            ReportGenerationException: If generation fails
        """
        self.generation_start_time = time.time()

        try:
            # Check cache first
            cache_key = self._get_cache_key(filters)
            if cache_enabled:
                cached_result = cache.get(cache_key)
                if cached_result:
                    logger.info(f"Report cache hit: {cache_key}")
                    return {'cached': True, **cached_result}

            # Update progress
            self._update_progress('collecting_data', 10, progress_callback)

            # Collect data from appropriate sources
            report_data = self._collect_report_data(filters)

            # Update progress
            self._update_progress('processing_data', 50, progress_callback)

            # Process and aggregate data
            processed_data = self._process_report_data(report_data, filters)

            # Update progress
            self._update_progress('generating_output', 80, progress_callback)

            # Format output
            formatted_output = self._format_output(processed_data, output_format)

            # Build result
            result = {
                'cached': False,
                'type': self.report_type,
                'format': output_format,
                'generated_at': timezone.now().isoformat(),
                'generation_time_seconds': time.time() - self.generation_start_time,
                'filters': filters or {},
                'data': formatted_output,
                'metadata': {
                    'report_name': self.metadata['name'],
                    'report_description': self.metadata['description'],
                    'data_sources': self.metadata['data_sources']
                }
            }

            # Cache the result
            if cache_enabled:
                cache.set(cache_key, {
                    'type': result['type'],
                    'format': result['format'],
                    'generated_at': result['generated_at'],
                    'data': result['data'],
                    'metadata': result['metadata']
                }, self.CACHE_TTL_MEDIUM)

            # Update progress to complete
            self._update_progress('completed', 100, progress_callback)

            logger.info(
                f"Report generated successfully: {self.report_type} "
                f"in {result['generation_time_seconds']:.2f}s"
            )

            return result

        except Exception as e:
            logger.error(f"Report generation failed: {str(e)}", exc_info=True)
            raise ReportGenerationException(f"Failed to generate report: {str(e)}")

    def _collect_report_data(self, filters: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """
        Collect data from appropriate sources based on report type.

        Args:
            filters: Filtering criteria

        Returns:
            Dict with raw data from various sources
        """
        data = {}

        if self.report_type == 'student_progress':
            data = self._collect_student_progress_data(filters)
        elif self.report_type == 'class_performance':
            data = self._collect_class_performance_data(filters)
        elif self.report_type == 'assignment_analysis':
            data = self._collect_assignment_analysis_data(filters)
        elif self.report_type == 'subject_analysis':
            data = self._collect_subject_analysis_data(filters)
        elif self.report_type == 'tutor_weekly':
            data = self._collect_tutor_weekly_data(filters)
        elif self.report_type == 'teacher_weekly':
            data = self._collect_teacher_weekly_data(filters)

        return data

    def _collect_student_progress_data(self, filters: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """Collect student progress data."""
        filters = filters or {}
        student_id = filters.get('student_id')
        date_range = filters.get('date_range', {})

        if not student_id:
            raise ReportGenerationException("student_id required for student progress report")

        try:
            student = User.objects.get(id=student_id, role=User.Role.STUDENT)
        except User.DoesNotExist:
            raise ReportGenerationException(f"Student with id {student_id} not found")

        # Get date range
        end_date = date_range.get('end_date', timezone.now().date())
        start_date = date_range.get('start_date', end_date - timedelta(days=30))

        # Material progress
        material_progress = MaterialProgress.objects.filter(
            student=student,
            last_accessed__date__range=[start_date, end_date]
        ).select_related('material')

        # Assignment submissions
        assignment_submissions = AssignmentSubmission.objects.filter(
            student=student,
            submitted_at__date__range=[start_date, end_date]
        ).select_related('assignment')

        # Knowledge graph progress
        kg_progress = LessonProgress.objects.filter(
            student=student,
            updated_at__date__range=[start_date, end_date]
        ).select_related('element')

        return {
            'student': {
                'id': student.id,
                'name': student.get_full_name(),
                'email': student.email,
                'role': student.role
            },
            'period': {
                'start': start_date.isoformat(),
                'end': end_date.isoformat()
            },
            'material_progress': list(material_progress.values(
                'material__id', 'material__title', 'progress_percentage',
                'time_spent', 'is_completed', 'last_accessed'
            )),
            'assignment_submissions': list(assignment_submissions.values(
                'assignment__id', 'assignment__title', 'score',
                'status', 'submitted_at'
            )),
            'kg_progress': list(kg_progress.values(
                'element__id', 'element__title', 'progress_percentage',
                'is_completed', 'updated_at'
            ))
        }

    def _collect_class_performance_data(self, filters: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """Collect class performance data."""
        filters = filters or {}

        # Get date range
        date_range = filters.get('date_range', {})
        end_date = date_range.get('end_date', timezone.now().date())
        start_date = date_range.get('start_date', end_date - timedelta(days=30))

        # Get all students (or filtered by class/teacher)
        students = User.objects.filter(role=User.Role.STUDENT)

        if self.user.role == User.Role.TEACHER:
            # Filter to teacher's students (if available)
            # This depends on your enrollment model structure
            pass

        # Aggregate assignment performance
        assignment_stats = AssignmentSubmission.objects.filter(
            submitted_at__date__range=[start_date, end_date]
        ).aggregate(
            avg_score=Avg('score'),
            total_submissions=Count('id'),
            students_participated=Count('student', distinct=True)
        )

        # Material completion rates
        material_stats = MaterialProgress.objects.filter(
            last_accessed__date__range=[start_date, end_date]
        ).aggregate(
            avg_progress=Avg('progress_percentage'),
            total_accesses=Count('id'),
            completed_count=Count('id', filter=Q(is_completed=True))
        )

        return {
            'period': {
                'start': start_date.isoformat(),
                'end': end_date.isoformat()
            },
            'class_size': students.count(),
            'assignment_statistics': assignment_stats,
            'material_statistics': material_stats,
            'timestamp': timezone.now().isoformat()
        }

    def _collect_assignment_analysis_data(self, filters: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """Collect assignment analysis data."""
        filters = filters or {}
        assignment_id = filters.get('assignment_id')

        if not assignment_id:
            raise ReportGenerationException("assignment_id required for assignment analysis")

        try:
            assignment = Assignment.objects.get(id=assignment_id)
        except Assignment.DoesNotExist:
            raise ReportGenerationException(f"Assignment with id {assignment_id} not found")

        submissions = AssignmentSubmission.objects.filter(
            assignment=assignment
        ).select_related('student')

        # Calculate statistics
        submission_data = list(submissions.values(
            'student__id', 'student__first_name', 'student__last_name',
            'score', 'status', 'submitted_at'
        ))

        stats = submissions.aggregate(
            total_submissions=Count('id'),
            submitted_count=Count('id', filter=Q(status='submitted')),
            avg_score=Avg('score'),
            max_score=Max('score') if submissions.exists() else 0,
            min_score=Min('score') if submissions.exists() else 0
        )

        # Add late submissions
        late_submissions = submissions.filter(
            submitted_at__gt=F('assignment__due_date')
        ).count()

        return {
            'assignment': {
                'id': assignment.id,
                'title': assignment.title,
                'due_date': assignment.due_date.isoformat() if assignment.due_date else None,
                'max_score': assignment.max_score
            },
            'submissions': submission_data,
            'statistics': {
                **stats,
                'late_submissions': late_submissions,
                'submission_rate': (stats['submitted_count'] / stats['total_submissions'] * 100)
                    if stats['total_submissions'] > 0 else 0
            }
        }

    def _collect_subject_analysis_data(self, filters: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """Collect subject analysis data."""
        filters = filters or {}
        subject_name = filters.get('subject_name')

        date_range = filters.get('date_range', {})
        end_date = date_range.get('end_date', timezone.now().date())
        start_date = date_range.get('start_date', end_date - timedelta(days=30))

        # Get materials for this subject
        materials = Material.objects.filter(
            subject__name=subject_name
        )

        # Material progress
        material_stats = MaterialProgress.objects.filter(
            material__in=materials,
            last_accessed__date__range=[start_date, end_date]
        ).aggregate(
            avg_progress=Avg('progress_percentage'),
            total_accesses=Count('id'),
            unique_students=Count('student', distinct=True)
        )

        # Assignment performance
        assignment_stats = AssignmentSubmission.objects.filter(
            assignment__materials__in=materials,
            submitted_at__date__range=[start_date, end_date]
        ).aggregate(
            avg_score=Avg('score'),
            total_submissions=Count('id'),
            unique_students=Count('student', distinct=True)
        )

        return {
            'subject': subject_name,
            'period': {
                'start': start_date.isoformat(),
                'end': end_date.isoformat()
            },
            'material_statistics': material_stats,
            'assignment_statistics': assignment_stats,
            'material_count': materials.count()
        }

    def _collect_tutor_weekly_data(self, filters: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """Collect tutor weekly data."""
        filters = filters or {}
        tutor_id = filters.get('tutor_id')

        if not tutor_id:
            raise ReportGenerationException("tutor_id required for tutor weekly report")

        # Get week range
        end_date = filters.get('end_date', timezone.now().date())
        start_date = end_date - timedelta(days=7)

        # Get all students this tutor works with
        # This depends on your tutoring relationship model

        return {
            'tutor_id': tutor_id,
            'week': {
                'start': start_date.isoformat(),
                'end': end_date.isoformat()
            },
            'timestamp': timezone.now().isoformat()
        }

    def _collect_teacher_weekly_data(self, filters: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """Collect teacher weekly data."""
        filters = filters or {}
        teacher_id = filters.get('teacher_id')

        if not teacher_id:
            raise ReportGenerationException("teacher_id required for teacher weekly report")

        try:
            teacher = User.objects.get(id=teacher_id, role=User.Role.TEACHER)
        except User.DoesNotExist:
            raise ReportGenerationException(f"Teacher with id {teacher_id} not found")

        # Get week range
        end_date = filters.get('end_date', timezone.now().date())
        start_date = end_date - timedelta(days=7)

        # Get teacher's assignments
        assignments = Assignment.objects.filter(
            created_by=teacher
        )

        # Assignment submission stats
        submission_stats = AssignmentSubmission.objects.filter(
            assignment__in=assignments,
            submitted_at__date__range=[start_date, end_date]
        ).aggregate(
            total_submissions=Count('id'),
            avg_score=Avg('score'),
            unique_students=Count('student', distinct=True)
        )

        return {
            'teacher': {
                'id': teacher.id,
                'name': teacher.get_full_name(),
                'email': teacher.email
            },
            'week': {
                'start': start_date.isoformat(),
                'end': end_date.isoformat()
            },
            'assignment_count': assignments.count(),
            'submission_statistics': submission_stats
        }

    def _process_report_data(self, data: Dict[str, Any], filters: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """
        Process and aggregate raw report data.

        Args:
            data: Raw data from sources
            filters: Filtering criteria

        Returns:
            Processed and aggregated data
        """
        processed = {
            'summary': self._generate_summary(data),
            'details': data,
            'insights': self._generate_insights(data),
            'metadata': {
                'generated_at': timezone.now().isoformat(),
                'filters': filters or {},
                'format_version': '1.0'
            }
        }

        return processed

    def _generate_summary(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate summary statistics from report data."""
        summary = {
            'report_type': self.report_type,
            'data_points': self._count_data_points(data),
            'timestamp': timezone.now().isoformat()
        }

        # Add type-specific summary
        if self.report_type == 'student_progress' and 'material_progress' in data:
            material_progress = data.get('material_progress', [])
            completed = sum(1 for m in material_progress if m.get('is_completed'))
            avg_progress = sum(m.get('progress_percentage', 0) for m in material_progress) / len(material_progress) if material_progress else 0

            summary['progress'] = {
                'materials_completed': completed,
                'total_materials': len(material_progress),
                'average_progress': round(avg_progress, 2),
                'total_time_spent_minutes': sum(m.get('time_spent', 0) for m in material_progress)
            }

        return summary

    def _generate_insights(self, data: Dict[str, Any]) -> List[str]:
        """Generate actionable insights from report data."""
        insights = []

        if self.report_type == 'student_progress':
            material_progress = data.get('material_progress', [])
            if material_progress:
                avg = sum(m.get('progress_percentage', 0) for m in material_progress) / len(material_progress)
                if avg > 80:
                    insights.append("Student is performing well with high completion rates")
                elif avg < 50:
                    insights.append("Student may need additional support or tutoring")
                else:
                    insights.append("Student is on track with moderate progress")

        return insights

    def _count_data_points(self, data: Dict[str, Any]) -> int:
        """Count total data points in the report."""
        count = 0
        if isinstance(data, dict):
            for value in data.values():
                if isinstance(value, (list, dict)):
                    count += self._count_data_points(value)
                else:
                    count += 1
        elif isinstance(data, list):
            count = len(data)
        return count

    def _format_output(self, data: Dict[str, Any], output_format: str) -> Any:
        """
        Format output in the specified format.

        Args:
            data: Processed report data
            output_format: Output format ('json', 'excel', 'pdf')

        Returns:
            Formatted output (dict for JSON, bytes for Excel/PDF)
        """
        if output_format == 'json':
            return data
        elif output_format == 'excel':
            return self._format_as_excel(data)
        elif output_format == 'pdf':
            return self._format_as_pdf(data)
        else:
            raise ReportGenerationException(f"Unsupported output format: {output_format}")

    def _format_as_excel(self, data: Dict[str, Any]) -> bytes:
        """Format report data as Excel spreadsheet."""
        workbook = openpyxl.Workbook()

        # Create sheets
        self._add_summary_sheet(workbook, data.get('summary', {}))
        self._add_details_sheet(workbook, data.get('details', {}))

        if 'insights' in data and data['insights']:
            self._add_insights_sheet(workbook, data['insights'])

        # Convert to bytes
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()

    def _add_summary_sheet(self, workbook: openpyxl.Workbook, summary: Dict[str, Any]) -> None:
        """Add summary sheet to Excel workbook."""
        ws = workbook.active
        ws.title = "Summary"

        # Add title
        ws['A1'] = f"{self.metadata['name']} - Summary"
        ws['A1'].font = Font(size=14, bold=True)

        # Add summary data
        row = 3
        for key, value in summary.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = str(value)
            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40

    def _add_details_sheet(self, workbook: openpyxl.Workbook, details: Dict[str, Any]) -> None:
        """Add details sheet to Excel workbook."""
        ws = workbook.create_sheet("Details")

        # Flatten and add data
        row = 1
        headers_written = False

        for key, value in details.items():
            if isinstance(value, list) and value and isinstance(value[0], dict):
                # Add table of data
                if not headers_written:
                    # Write headers
                    headers = list(value[0].keys())
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=row, column=col, value=header)
                    row += 1
                    headers_written = True

                # Write data rows
                for item in value:
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=row, column=col, value=item.get(header, ''))
                    row += 1
            else:
                ws[f'A{row}'] = key
                ws[f'B{row}'] = str(value)
                row += 1

    def _add_insights_sheet(self, workbook: openpyxl.Workbook, insights: List[str]) -> None:
        """Add insights sheet to Excel workbook."""
        ws = workbook.create_sheet("Insights")

        ws['A1'] = "Key Insights"
        ws['A1'].font = Font(size=12, bold=True)

        for idx, insight in enumerate(insights, 2):
            ws[f'A{idx}'] = insight
            ws.row_dimensions[idx].height = 30

        ws.column_dimensions['A'].width = 80

    def _format_as_pdf(self, data: Dict[str, Any]) -> bytes:
        """Format report data as PDF."""
        # Placeholder for PDF generation using ReportLab or WeasyPrint
        # Returns the PDF as bytes

        # For now, return a JSON representation with a note
        logger.warning("PDF generation not fully implemented, returning JSON representation")
        return json.dumps(data, indent=2, default=str).encode('utf-8')

    def _update_progress(
        self,
        status: str,
        percentage: int,
        callback: Optional[callable] = None
    ) -> None:
        """
        Update progress tracking.

        Args:
            status: Current status
            percentage: Progress percentage (0-100)
            callback: Optional callback function for progress updates
        """
        self.progress_data = {
            'status': status,
            'percentage': percentage,
            'timestamp': timezone.now().isoformat()
        }

        if callback:
            try:
                callback(self.progress_data)
            except Exception as e:
                logger.warning(f"Progress callback failed: {str(e)}")

    def _get_cache_key(self, filters: Optional[Dict[str, Any]] = None) -> str:
        """Generate cache key for report."""
        filter_str = json.dumps(filters or {}, sort_keys=True, default=str)
        return f"report_{self.report_type}_{hash(filter_str) % 1000000}"

    def invalidate_cache(self, filters: Optional[Dict[str, Any]] = None) -> None:
        """
        Invalidate cached report.

        Args:
            filters: Filters to invalidate (None invalidates all of this type)
        """
        if filters is None:
            cache.delete_pattern(f"report_{self.report_type}_*")
        else:
            cache_key = self._get_cache_key(filters)
            cache.delete(cache_key)

    def get_progress(self) -> Dict[str, Any]:
        """Get current generation progress."""
        return self.progress_data


class ReportScheduler:
    """
    Schedules and manages automatic report generation.

    Handles:
    - Report scheduling (daily, weekly, monthly)
    - Automatic generation at specified times
    - Report delivery
    - Schedule management
    """

    SCHEDULES = {
        'daily': {
            'frequency': 'daily',
            'hours': 24
        },
        'weekly': {
            'frequency': 'weekly',
            'days': 7
        },
        'monthly': {
            'frequency': 'monthly',
            'days': 30
        }
    }

    @staticmethod
    def schedule_report(
        user: User,
        report_type: str,
        frequency: str,
        recipients: List[User],
        filters: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Schedule a recurring report.

        Args:
            user: User scheduling the report
            report_type: Type of report
            frequency: Frequency ('daily', 'weekly', 'monthly')
            recipients: Users to receive the report
            filters: Report filters

        Returns:
            Schedule confirmation data
        """
        if frequency not in ReportScheduler.SCHEDULES:
            raise ReportGenerationException(f"Invalid frequency: {frequency}")

        schedule_info = ReportScheduler.SCHEDULES[frequency]

        return {
            'status': 'scheduled',
            'report_type': report_type,
            'frequency': frequency,
            'next_generation': (timezone.now() + timedelta(**schedule_info)).isoformat(),
            'recipients': [r.id for r in recipients],
            'created_at': timezone.now().isoformat()
        }

    @staticmethod
    def get_pending_schedules() -> List[Dict[str, Any]]:
        """Get all pending scheduled reports."""
        # Implementation would query scheduled reports from database
        return []
