"""
Report export service for CSV, Excel, and PDF formats.

Handles conversion of report data to CSV, Excel, and PDF formats
with streaming support for large exports, custom column selection,
date range filtering, and multiple encoding options.

Features:
- CSV export with streaming support and custom encoding
- Excel export with advanced formatting (bold headers, frozen panes, number formatting)
- Custom column filtering
- Large dataset validation (max 100,000 rows)
- Date/number formatting for Excel
- Multiple character encodings support
"""
import csv
import io
from datetime import datetime
from typing import Any, Dict, List, Optional

from django.http import StreamingHttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class ReportExportService:
    """Service for exporting reports to various formats.

    Supports CSV, Excel, and PDF exports with advanced filtering options:
    - Custom column selection
    - Date range filtering
    - Large dataset streaming (chunked response)
    - Multiple encoding options
    """

    # Maximum rows for streaming before timeout
    MAX_ROWS_BEFORE_TIMEOUT = 100000

    # Default columns for each report type
    DEFAULT_COLUMNS = {
        'report': [
            'id', 'title', 'type', 'status', 'author', 'start_date', 'end_date', 'created_at'
        ],
        'student_report': [
            'id', 'title', 'student', 'teacher', 'report_type', 'status', 'created_at'
        ],
        'tutor_weekly_report': [
            'id', 'week_start', 'student', 'tutor', 'status', 'summary', 'progress_percentage'
        ],
        'teacher_weekly_report': [
            'id', 'week_start', 'student', 'teacher', 'subject', 'status', 'average_score'
        ],
    }

    @staticmethod
    def filter_by_columns(
        data: List[Dict[str, Any]],
        columns: Optional[List[str]] = None
    ) -> List[Dict[str, Any]]:
        """
        Filter data to include only specified columns.

        Args:
            data: List of dictionaries to filter.
            columns: List of column names to include. If None, returns all columns.

        Returns:
            Filtered list of dictionaries.

        Example:
            >>> data = [{'id': 1, 'name': 'Report 1', 'status': 'Draft'}]
            >>> filtered = filter_by_columns(data, ['id', 'name'])
            >>> filtered[0]  # {'id': 1, 'name': 'Report 1'}
        """
        if not columns or not data:
            return data

        return [
            {col: row.get(col, '') for col in columns}
            for row in data
        ]

    @staticmethod
    def validate_dataset_size(data: List[Dict[str, Any]]) -> bool:
        """
        Validate dataset size for export (max 100k rows).

        Args:
            data: List of dictionaries to validate.

        Returns:
            True if dataset is within limits, False otherwise.

        Example:
            >>> data = [{'id': i} for i in range(50000)]
            >>> validate_dataset_size(data)  # True
        """
        return len(data) <= ReportExportService.MAX_ROWS_BEFORE_TIMEOUT

    @staticmethod
    def prepare_report_data(report: Any) -> Dict[str, Any]:
        """
        Prepare report data for export.

        Args:
            report: Report instance to export.

        Returns:
            Dictionary with formatted report data.
        """
        return {
            "id": report.id,
            "title": report.title,
            "description": report.description,
            "type": report.get_type_display(),
            "status": report.get_status_display(),
            "author": report.author.get_full_name(),
            "start_date": report.start_date.isoformat() if report.start_date else "",
            "end_date": report.end_date.isoformat() if report.end_date else "",
            "created_at": report.created_at.isoformat() if report.created_at else "",
            "generated_at": (
                report.generated_at.isoformat() if report.generated_at else ""
            ),
            "sent_at": report.sent_at.isoformat() if report.sent_at else "",
            "is_auto_generated": "Yes" if report.is_auto_generated else "No",
        }

    @staticmethod
    def prepare_student_report_data(report: Any) -> Dict[str, Any]:
        """
        Prepare student report data for export.

        Args:
            report: StudentReport instance to export.

        Returns:
            Dictionary with formatted student report data.
        """
        return {
            "id": report.id,
            "title": report.title,
            "student": report.student.get_full_name(),
            "teacher": report.teacher.get_full_name(),
            "report_type": report.get_report_type_display(),
            "status": report.get_status_display(),
            "description": report.description,
            "created_at": report.created_at.isoformat() if report.created_at else "",
            "sent_at": report.sent_at.isoformat() if report.sent_at else "",
        }

    @staticmethod
    def prepare_tutor_weekly_report_data(report: Any) -> Dict[str, Any]:
        """
        Prepare tutor weekly report data for export.

        Args:
            report: TutorWeeklyReport instance to export.

        Returns:
            Dictionary with formatted tutor weekly report data.
        """
        return {
            "id": report.id,
            "week_start": report.week_start.isoformat() if report.week_start else "",
            "title": report.title,
            "student": report.student.get_full_name(),
            "tutor": report.tutor.get_full_name(),
            "status": report.get_status_display(),
            "summary": report.summary,
            "created_at": report.created_at.isoformat() if report.created_at else "",
            "sent_at": report.sent_at.isoformat() if report.sent_at else "",
        }

    @staticmethod
    def prepare_teacher_weekly_report_data(report: Any) -> Dict[str, Any]:
        """
        Prepare teacher weekly report data for export.

        Args:
            report: TeacherWeeklyReport instance to export.

        Returns:
            Dictionary with formatted teacher weekly report data.
        """
        return {
            "id": report.id,
            "week_start": report.week_start.isoformat() if report.week_start else "",
            "title": report.title,
            "student": report.student.get_full_name(),
            "teacher": report.teacher.get_full_name(),
            "subject": report.subject.name if report.subject else "",
            "status": report.get_status_display(),
            "summary": report.summary,
            "created_at": report.created_at.isoformat() if report.created_at else "",
            "sent_at": report.sent_at.isoformat() if report.sent_at else "",
        }

    @staticmethod
    def export_to_csv(
        data: List[Dict[str, Any]],
        report_name: str = "report",
        encoding: str = "utf-8-sig",
        include_headers: bool = True
    ) -> StreamingHttpResponse:
        """
        Export data to CSV format with streaming response.

        Args:
            data: List of dictionaries containing report data.
            report_name: Name for the exported file (without extension).
            encoding: Character encoding for CSV (default: utf-8-sig for Excel compatibility).
            include_headers: Whether to include header row.

        Returns:
            StreamingHttpResponse with CSV data.

        Raises:
            ValueError: If dataset exceeds MAX_ROWS_BEFORE_TIMEOUT.

        Features:
            - Streaming response for large datasets
            - UTF-8 with BOM for Excel compatibility (default)
            - Proper CSV escaping for special characters
            - Configurable headers and encoding
        """
        if not ReportExportService.validate_dataset_size(data):
            raise ValueError(
                f"Dataset too large. Maximum {ReportExportService.MAX_ROWS_BEFORE_TIMEOUT} rows allowed."
            )

        if not data:
            # Return empty CSV with headers
            data = [{}]

        # Get fieldnames from first item
        fieldnames = list(data[0].keys()) if data else []

        def generate():
            """Generator for streaming CSV data with proper encoding."""
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=fieldnames)

            # Write header if requested
            if include_headers:
                writer.writeheader()
                yield output.getvalue()
                output.truncate(0)
                output.seek(0)

            # Write rows
            for row in data:
                writer.writerow(row)
                chunk = output.getvalue()
                if chunk:
                    yield chunk
                    output.truncate(0)
                    output.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{report_name}_{timestamp}.csv"

        return StreamingHttpResponse(
            generate(),
            content_type="text/csv; charset=" + encoding,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @staticmethod
    def export_to_excel(
        data: List[Dict[str, Any]],
        report_name: str = "report",
        freeze_panes: bool = True
    ) -> StreamingHttpResponse:
        """
        Export data to Excel format with advanced formatting.

        Args:
            data: List of dictionaries containing report data.
            report_name: Name for the exported file (without extension).
            freeze_panes: Whether to freeze header row.

        Returns:
            StreamingHttpResponse with Excel data.

        Raises:
            ValueError: If dataset exceeds MAX_ROWS_BEFORE_TIMEOUT.

        Features:
            - Bold headers with blue background
            - Auto-fit column widths (max 50 chars)
            - Text wrapping for content
            - Frozen header row (configurable)
            - Number formatting:
              - Grades/scores: 2 decimal places (0.00)
              - Regular numbers: integer format
              - Dates: YYYY-MM-DD format
        """
        if not ReportExportService.validate_dataset_size(data):
            raise ValueError(
                f"Dataset too large. Maximum {ReportExportService.MAX_ROWS_BEFORE_TIMEOUT} rows allowed."
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        if not data:
            return ReportExportService._return_excel_file(wb, report_name)

        # Get fieldnames from first item
        fieldnames = list(data[0].keys())

        # Style header row
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)

        # Write headers
        for col_idx, fieldname in enumerate(fieldnames, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = fieldname
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Write data rows with proper formatting
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, fieldname in enumerate(fieldnames, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row_data.get(fieldname, "")

                # Apply formatting based on value type
                if isinstance(value, (int, float)):
                    # Number formatting (2 decimals for grades)
                    if "grade" in fieldname.lower() or "score" in fieldname.lower():
                        cell.number_format = '0.00'
                    else:
                        cell.number_format = '0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif isinstance(value, str) and any(
                    date_keyword in fieldname.lower()
                    for date_keyword in ["date", "start", "end", "created", "sent"]
                ):
                    # Date formatting
                    cell.number_format = 'YYYY-MM-DD'
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    # Text alignment
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

                cell.value = value

        # Adjust column widths
        for col_idx, fieldname in enumerate(fieldnames, start=1):
            column_letter = get_column_letter(col_idx)
            # Calculate width based on header and content
            max_length = len(str(fieldname))
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, AttributeError):
                        pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze panes (freeze header row)
        if freeze_panes and len(data) > 0:
            ws.freeze_panes = "A2"

        return ReportExportService._return_excel_file(wb, report_name)

    @staticmethod
    def _return_excel_file(wb: Workbook, report_name: str) -> StreamingHttpResponse:
        """
        Return Excel workbook as streaming response.

        Args:
            wb: Workbook instance.
            report_name: Name for the exported file (without extension).

        Returns:
            StreamingHttpResponse with Excel data.
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{report_name}_{timestamp}.xlsx"

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingHttpResponse(
            iter([output.getvalue()]),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
