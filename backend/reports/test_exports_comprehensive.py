"""
Comprehensive tests for CSV and Excel export functionality for reports.

Tests cover:
- CSV export with various options (encoding, headers, etc.)
- Excel export with advanced formatting
- Column filtering
- Dataset size validation
- Large dataset streaming
- Special character handling
- Permission and authentication checks
"""
import csv
import io
from datetime import date, datetime, timedelta

import pytest
from django.contrib.auth import get_user_model
from openpyxl import load_workbook
from rest_framework.test import APIClient, APITestCase

from accounts.models import StudentProfile
from materials.models import Subject
from reports.models import (
    Report,
    StudentReport,
    TeacherWeeklyReport,
    TutorWeeklyReport,
)
from reports.services.export import ReportExportService

User = get_user_model()


def get_streaming_content(response):
    """Helper to get content from StreamingHttpResponse."""
    if hasattr(response, "streaming_content"):
        # StreamingHttpResponse
        return b"".join(response.streaming_content)
    else:
        # Regular HttpResponse
        return response.content


class BasicExportServiceTests(APITestCase):
    """Test ReportExportService core functionality."""

    def setUp(self):
        """Set up test data."""
        self.service = ReportExportService

    def test_filter_by_columns_with_columns(self):
        """Test filtering data to specific columns."""
        data = [
            {"id": 1, "name": "Report 1", "status": "Draft", "author": "John"},
            {"id": 2, "name": "Report 2", "status": "Sent", "author": "Jane"},
        ]

        filtered = self.service.filter_by_columns(data, columns=["id", "name"])

        assert len(filtered) == 2
        assert "id" in filtered[0]
        assert "name" in filtered[0]
        assert "status" not in filtered[0]
        assert "author" not in filtered[0]
        assert filtered[0]["id"] == 1
        assert filtered[0]["name"] == "Report 1"

    def test_filter_by_columns_without_columns(self):
        """Test that all columns are returned if none specified."""
        data = [{"id": 1, "name": "Report 1", "status": "Draft"}]

        filtered = self.service.filter_by_columns(data, columns=None)

        assert len(filtered[0]) == 3
        assert all(key in filtered[0] for key in ["id", "name", "status"])

    def test_filter_by_columns_empty_data(self):
        """Test filtering empty data returns empty."""
        result = self.service.filter_by_columns([], columns=["id"])

        assert result == []

    def test_validate_dataset_size_under_limit(self):
        """Test validation for dataset under limit."""
        data = [{"id": i, "name": f"Report {i}"} for i in range(1000)]

        assert self.service.validate_dataset_size(data) is True

    def test_validate_dataset_size_at_limit(self):
        """Test validation for dataset at exact limit."""
        data = [{"id": i} for i in range(100000)]

        assert self.service.validate_dataset_size(data) is True

    def test_validate_dataset_size_over_limit(self):
        """Test validation for dataset over limit."""
        data = [{"id": i} for i in range(100001)]

        assert self.service.validate_dataset_size(data) is False

    def test_export_csv_dataset_too_large(self):
        """Test that CSV export raises error for dataset over limit."""
        data = [{"id": i} for i in range(100001)]

        with pytest.raises(ValueError, match="Dataset too large"):
            self.service.export_to_csv(data)

    def test_export_excel_dataset_too_large(self):
        """Test that Excel export raises error for dataset over limit."""
        data = [{"id": i} for i in range(100001)]

        with pytest.raises(ValueError, match="Dataset too large"):
            self.service.export_to_excel(data)


class CSVExportTests(APITestCase):
    """Test CSV export functionality."""

    def setUp(self):
        """Set up test data."""
        self.service = ReportExportService

    def test_export_to_csv_empty(self):
        """Test exporting empty data to CSV."""
        response = self.service.export_to_csv([])

        assert response.status_code == 200
        assert response["Content-Type"] == "text/csv; charset=utf-8-sig"
        assert "attachment" in response["Content-Disposition"]

    def test_export_to_csv_with_data(self):
        """Test exporting data to CSV with actual content."""
        data = [
            {"id": 1, "name": "Report 1", "status": "Draft"},
            {"id": 2, "name": "Report 2", "status": "Sent"},
        ]

        response = self.service.export_to_csv(data, "test_report")

        assert response.status_code == 200
        content = get_streaming_content(response).decode("utf-8")
        csv_reader = csv.DictReader(io.StringIO(content))
        rows = list(csv_reader)

        assert len(rows) == 2
        assert rows[0]["name"] == "Report 1"
        assert rows[1]["status"] == "Sent"

    def test_export_csv_with_custom_encoding(self):
        """Test CSV export with different encoding."""
        data = [
            {"id": 1, "name": "Тест", "status": "Draft"},
            {"id": 2, "name": "中文", "status": "Sent"},
        ]

        response = self.service.export_to_csv(data, encoding="utf-8-sig")

        assert response.status_code == 200
        assert "charset=utf-8-sig" in response["Content-Type"]

    def test_export_csv_with_headers(self):
        """Test CSV export includes headers."""
        data = [{"id": 1, "name": "Report 1"}]

        response = self.service.export_to_csv(data, include_headers=True)

        content = get_streaming_content(response).decode("utf-8")
        lines = content.strip().split("\n")

        assert len(lines) == 2
        assert "id" in lines[0]
        assert "name" in lines[0]

    def test_export_csv_without_headers(self):
        """Test CSV export without header row."""
        data = [
            {"id": 1, "name": "Report 1"},
            {"id": 2, "name": "Report 2"},
        ]

        response = self.service.export_to_csv(data, include_headers=False)

        content = get_streaming_content(response).decode("utf-8")
        lines = [line for line in content.strip().split("\n") if line]

        assert len(lines) == 2  # Only data rows

    def test_export_csv_special_characters(self):
        """Test CSV export handles special characters properly."""
        data = [
            {
                "id": 1,
                "name": 'Report "with quotes"',
                "description": "Line 1\nLine 2"
            },
            {"id": 2, "name": "Report with, comma", "description": "Normal text"},
        ]

        response = self.service.export_to_csv(data)

        content = get_streaming_content(response).decode("utf-8")
        csv_reader = csv.DictReader(io.StringIO(content))
        rows = list(csv_reader)

        assert len(rows) == 2
        assert "quotes" in rows[0]["name"]
        assert "Line 2" in rows[0]["description"]

    def test_csv_filename_includes_timestamp(self):
        """Test that CSV filename includes timestamp."""
        response = self.service.export_to_csv([], "test_report")

        disposition = response["Content-Disposition"]
        assert "test_report_" in disposition
        assert ".csv" in disposition
        assert "attachment" in disposition


class ExcelExportTests(APITestCase):
    """Test Excel export functionality."""

    def setUp(self):
        """Set up test data."""
        self.service = ReportExportService

    def test_export_to_excel_empty(self):
        """Test exporting empty data to Excel."""
        response = self.service.export_to_excel([])

        assert response.status_code == 200
        assert "spreadsheetml" in response["Content-Type"]

    def test_export_to_excel_with_data(self):
        """Test exporting data to Excel with actual content."""
        data = [
            {"id": 1, "name": "Report 1", "status": "Draft"},
            {"id": 2, "name": "Report 2", "status": "Sent"},
        ]

        response = self.service.export_to_excel(data, "test_report")

        assert response.status_code == 200
        content = get_streaming_content(response)
        workbook = load_workbook(io.BytesIO(content))
        worksheet = workbook.active

        assert worksheet["A1"].value == "id"
        assert worksheet["B1"].value == "name"
        assert worksheet["A2"].value == 1
        assert worksheet["B2"].value == "Report 1"

    def test_export_excel_with_freeze_panes(self):
        """Test Excel export with frozen header row."""
        data = [
            {"id": 1, "name": "Report 1"},
            {"id": 2, "name": "Report 2"},
        ]

        response = self.service.export_to_excel(data, freeze_panes=True)

        content = get_streaming_content(response)
        workbook = load_workbook(io.BytesIO(content))
        worksheet = workbook.active

        assert worksheet.freeze_panes == "A2"

    def test_export_excel_without_freeze_panes(self):
        """Test Excel export without frozen header row."""
        data = [{"id": 1, "name": "Report 1"}]

        response = self.service.export_to_excel(data, freeze_panes=False)

        content = get_streaming_content(response)
        workbook = load_workbook(io.BytesIO(content))
        worksheet = workbook.active

        assert worksheet.freeze_panes is None

    def test_excel_header_styling(self):
        """Test Excel header has proper styling (bold, blue background)."""
        data = [{"name": "Test", "value": 123}]

        response = self.service.export_to_excel(data)

        content = get_streaming_content(response)
        workbook = load_workbook(io.BytesIO(content))
        worksheet = workbook.active

        header_cell = worksheet["A1"]
        assert header_cell.font.bold is True
        assert header_cell.font.color.rgb is not None
        assert header_cell.fill.start_color.rgb is not None

    def test_excel_number_formatting_for_grades(self):
        """Test that grades are formatted to 2 decimal places in Excel."""
        data = [
            {"id": 1, "grade": 95.5, "score": 87.666},
            {"id": 2, "grade": 92.1, "score": 91.234},
        ]

        response = self.service.export_to_excel(data)

        content = get_streaming_content(response)
        workbook = load_workbook(io.BytesIO(content))
        worksheet = workbook.active

        # Check grade column formatting (should be 0.00)
        grade_cell = worksheet["B2"]
        assert grade_cell.number_format == "0.00"

    def test_excel_number_formatting_for_integers(self):
        """Test that integer numbers are formatted without decimals."""
        data = [
            {"id": 1, "count": 42},
            {"id": 2, "count": 100},
        ]

        response = self.service.export_to_excel(data)

        content = get_streaming_content(response)
        workbook = load_workbook(io.BytesIO(content))
        worksheet = workbook.active

        count_cell = worksheet["B2"]
        assert count_cell.number_format == "0"

    def test_excel_auto_fit_column_widths(self):
        """Test that Excel columns are auto-fitted."""
        data = [{"short": "a", "very_long_column_name": "very long content here"}]

        response = self.service.export_to_excel(data)

        content = get_streaming_content(response)
        workbook = load_workbook(io.BytesIO(content))
        worksheet = workbook.active

        # Long column should be wider than short column
        short_width = worksheet.column_dimensions["A"].width
        long_width = worksheet.column_dimensions["B"].width

        assert long_width > short_width

    def test_excel_filename_includes_timestamp(self):
        """Test that Excel filename includes timestamp."""
        response = self.service.export_to_excel([], "test_report")

        disposition = response["Content-Disposition"]
        assert "test_report_" in disposition
        assert ".xlsx" in disposition


class ReportExportAPITests(APITestCase):
    """Test report export through API endpoints."""

    def setUp(self):
        """Set up test data."""
        self.client = APIClient()

        # Create users
        self.teacher = User.objects.create_user(
            username="teacher",
            email="teacher@test.com",
            password="TestPass123!",
            first_name="John",
            last_name="Teacher",
            role="teacher",
        )

        self.student1 = User.objects.create_user(
            username="student1",
            email="student1@test.com",
            password="TestPass123!",
            first_name="Alice",
            last_name="Student",
            role="student",
        )

        # Create reports
        self.report1 = Report.objects.create(
            title="Report 1",
            description="Test report 1",
            type=Report.Type.STUDENT_PROGRESS,
            status=Report.Status.DRAFT,
            author=self.teacher,
            start_date=date.today() - timedelta(days=7),
            end_date=date.today(),
        )
        self.report1.target_students.add(self.student1)

    def test_export_reports_csv(self):
        """Test exporting reports to CSV format through API."""
        self.client.force_authenticate(user=self.teacher)
        response = self.client.get("/api/reports/export/?format=csv")

        assert response.status_code == 200
        assert response["Content-Type"].startswith("text/csv")

    def test_export_reports_excel(self):
        """Test exporting reports to Excel format through API."""
        self.client.force_authenticate(user=self.teacher)
        response = self.client.get("/api/reports/export/?format=xlsx")

        assert response.status_code == 200
        assert "spreadsheetml" in response["Content-Type"]

    def test_export_invalid_format(self):
        """Test exporting with invalid format parameter."""
        self.client.force_authenticate(user=self.teacher)
        response = self.client.get("/api/reports/export/?format=invalid")

        assert response.status_code == 400
        assert "Invalid format" in response.json()["error"]

    def test_export_requires_authentication(self):
        """Test that export requires authentication."""
        response = self.client.get("/api/reports/export/?format=csv")

        assert response.status_code == 401


class LargeDatasetExportTests(APITestCase):
    """Test export with large datasets."""

    def setUp(self):
        """Set up test data."""
        self.service = ReportExportService

    def test_export_large_dataset_streaming(self):
        """Test streaming export of large dataset (under limit)."""
        data = [
            {"id": i, "name": f"Report {i}", "status": "Draft"}
            for i in range(10000)
        ]

        response = self.service.export_to_csv(data)

        assert response.status_code == 200
        assert response.streaming

    def test_csv_export_data_integrity_large(self):
        """Test CSV export maintains data integrity with large dataset."""
        data = [
            {"id": i, "value": 100 + i, "percentage": 99.99 - (i * 0.01)}
            for i in range(1000)
        ]

        response = self.service.export_to_csv(data)

        content = get_streaming_content(response).decode("utf-8")
        csv_reader = csv.DictReader(io.StringIO(content))
        rows = list(csv_reader)

        assert len(rows) == 1000
        assert float(rows[0]["percentage"]) == 99.99
        assert int(rows[999]["value"]) == 1099

    def test_excel_export_multiple_sheets(self):
        """Test Excel export creates single sheet with all data."""
        data = [{"id": i, "name": f"Report {i}"} for i in range(100)]

        response = self.service.export_to_excel(data)

        content = get_streaming_content(response)
        workbook = load_workbook(io.BytesIO(content))

        assert len(workbook.sheetnames) == 1
        assert workbook.sheetnames[0] == "Report"


class DataIntegrityTests(APITestCase):
    """Test data integrity in exports."""

    def setUp(self):
        """Set up test data."""
        self.service = ReportExportService

    def test_csv_special_characters_preserved(self):
        """Test special characters are preserved in CSV."""
        special_data = [
            {"name": "Test™", "description": "Contains © symbol"},
            {"name": "Über", "description": "With ü umlaut"},
        ]

        response = self.service.export_to_csv(special_data)

        content = get_streaming_content(response).decode("utf-8")
        assert "™" in content
        assert "©" in content
        assert "Über" in content

    def test_csv_numeric_precision(self):
        """Test numeric precision is preserved in CSV."""
        data = [
            {"grade": 95.555, "percentage": 99.999},
        ]

        response = self.service.export_to_csv(data)

        content = get_streaming_content(response).decode("utf-8")
        assert "95.555" in content
        assert "99.999" in content

    def test_excel_text_wrapping_enabled(self):
        """Test text wrapping is enabled in Excel."""
        data = [
            {
                "description": "This is a very long text that should wrap in Excel"
            },
        ]

        response = self.service.export_to_excel(data)

        content = get_streaming_content(response)
        workbook = load_workbook(io.BytesIO(content))
        worksheet = workbook.active

        data_cell = worksheet["A2"]
        assert data_cell.alignment.wrap_text is True
